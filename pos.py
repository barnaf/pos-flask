from flask import Flask, render_template, request
import openpyxl
from datetime import datetime
import os

app = Flask(__name__)
FILE_NAME = 'pos_data.xlsx'

# تحميل أو إنشاء ملف Excel
if not os.path.exists(FILE_NAME):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    inv = wb.create_sheet('Inventory')
    sales = wb.create_sheet('Sales')
    inv.append(['ProductCode', 'ProductName', 'Price', 'Quantity'])
    sales.append(['DateTime', 'ProductCode', 'ProductName', 'Quantity', 'PaymentMethod', 'Expense'])
    # مثال منتجات
    inv.append(['P001', 'عطر الشوق', 100, 10])
    inv.append(['P002', 'عطر الليل', 150, 5])
    inv.append(['P003', 'عطر الورد', 120, 8])
    wb.save(FILE_NAME)
else:
    wb = openpyxl.load_workbook(FILE_NAME)
    inv = wb['Inventory']
    sales = wb['Sales']

# دالة لجلب المنتجات للموقع
def get_products():
    products = []
    for row in inv.iter_rows(min_row=2, values_only=True):
        products.append({
            'code': row[0],
            'name': row[1],
            'price': row[2],
            'quantity': row[3]
        })
    return products

# الصفحة الرئيسية - عرض المنتجات
@app.route('/')
def home():
    products = get_products()
    return render_template('home.html', products=products)

# دالة البيع - هذه اللي تسوي البيع فعليًا
@app.route('/sell', methods=['POST'])
def sell():
    code = request.form['code']
    quantity = int(request.form['quantity'])
    payment = request.form['payment']
    expense = float(request.form.get('expense', 0))

    print("تحديث المخزون...")
    for row in inv.iter_rows(min_row=2):
        if row[0].value == code:
            if row[3].value < quantity:
                return "الكمية المطلوبة أكبر من المخزون!", 400
            row[3].value -= quantity
            product_name = row[1].value
            product_price = row[2].value
            break
    else:
        return "المنتج غير موجود", 400

    print("إضافة سجل البيع...")
    sales.append([datetime.now().strftime('%Y-%m-%d %H:%M:%S'), code, product_name, quantity, payment, expense])

    print("حفظ الملف...")
    wb.save(FILE_NAME)
    print("تم الحفظ بنجاح")

    total_price = product_price * quantity

    return render_template('receipt.html',
                           product_name=product_name,
                           quantity=quantity,
                           total_price=total_price,
                           payment=payment,
                           date=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                           expense=expense)

if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True)
