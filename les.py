from openpyxl import Workbook, load_workbook

#wb = load_workbook ()

#wb = workbook()import openpyxl

wb = Workbook()           # إنشاء ملف Excel جديد
ws = wb.active            # الوصول للورقة الافتراضية الأولى

ws.title = 'New title'     # ✅ صحيح
  # تغيير اسم الورقة
print(wb.sheetnames)
ws['C2'] = 12


# اضافة صفوف كاملة  
ws.append(["hello","python", "pythonandExls", 1,2,3])

#cell = ws ['C2']
#print (cell.value)

colm = ws ['C']
print (colm)

row_3 = ws [2]
print (row_3)

cell_range = ws ['A3':'F3']
print (cell_range)
for cell in cell_range[0]:
    print(cell.value)

ws.merge_cells ("A1:D3")
ws["A1 "] = 23



wb.save("Myworkspace.xlsx")  # حفظ الملف


